""" 
把 excel 导出 json | lua
"""

import json
import os
import time

from openpyxl import load_workbook
from openpyxl.worksheet import worksheet

import cfglistjson
import configts
import json2lua
import sheetindex
import str2list


class SheetStruct:
    """ 表的导出信息 """

    # 服务端导出文件名称
    def serverName(self):
        pass

    # 客户端导出文件名称
    def clientName(self):
        pass

    # key的数量
    def keyCount(self):
        pass

    # 特殊的格式
    def spcialType(self):
        pass

    # sheet名称
    def sheetTitle(self):
        pass

    # xlsx名称
    def xlsxTitle(self):
        pass


class DataStruct:
    """ 导出数据的结构体信息 """

    # 字段名
    def _name(self):
        pass

    # 类型 (number, string, array, object)
    def _type(self):
        pass

    # 导出字段（S服务端C客户端）
    def _cs(self):
        pass

    # 字段名注释
    def _def(self):
        pass


class Excel2Json:
    xlslUrl = ''
    # 配表信息定义的行数
    structRow = [4, 5, 6, 7]
    # 配表真实数据要导出的列数，每行的后面可能有些说明，但是是不用导出，所以需要记录要导出的真实列数，也就是structRow行的真实列数
    structColLen = 0
    # 配表真实数据开始的行数
    startRow = 8
    sheet: worksheet.Worksheet = None
    # 表头结构体
    sheetStruct: SheetStruct = None
    # 导出路径
    outputRoot: str = ''

    def __init__(self, xlsx_url: str, output_root: str) -> None:
        self.xlslUrl = xlsx_url
        self.outputRoot = output_root

    def read_file(self) -> None:
        xlsxTitle = self.get_xlsx_title()
        print('===== 开始处理表：', xlsxTitle, " =====")
        wb = load_workbook(filename=self.xlslUrl)
        # print(wb.sheetnames)
        # print(wb.worksheets)
        for sheet in wb.worksheets:
            # title中以#开头的表示不导出
            if sheet.title[0] == '#':
                continue
            self.sheet = sheet
            if not self.get_sheet_struct():
                continue
            self.deal_single_sheet()
        self.sheet = None
        self.xlslUrl = ''
        self.sheetStruct = None
        print('===== 结束处理', xlsxTitle, " =====\n")

    def get_xlsx_title(self) -> str:
        """ 获取xlsx文件名称 """
        basename = os.path.basename(self.xlslUrl).replace('.xlsx', '')
        return basename

    def get_sheet_struct(self) -> SheetStruct:
        """ 获取当个sheet导出文件的名称以及key数量 """
        serverName = self.sheet.cell(row=1, column=2).value
        clientName = self.sheet.cell(row=1, column=5).value
        keyCount = self.sheet.cell(row=2, column=2).value
        specialType = self.sheet.cell(row=2, column=5).value
        if not (serverName or clientName):
            return None
        struct = SheetStruct()
        struct.serverName = serverName
        struct.clientName = clientName
        struct.keyCount = keyCount
        struct.spcialType = specialType == 1
        struct.sheetTitle = self.sheet.title
        struct.xlsxTitle = self.get_xlsx_title()
        self.sheetStruct = struct
        return struct

    def get_row_value(self, row: int) -> list:
        """ 获取某行的数据 """
        columns = self.sheet.max_column
        rowData = []
        for i in range(1, columns + 1):
            cellValue = self.sheet.cell(row=row, column=i).value
            # 单元格填0是需要导出的
            if cellValue is None:
                break
            rowData.append(cellValue)
        return rowData

    def get_data_struct(self) -> dict:
        """ 导出数据的结构体信息 """
        row4 = self.get_row_value(self.structRow[0])
        row5 = self.get_row_value(self.structRow[1])
        row6 = self.get_row_value(self.structRow[2])
        row7 = self.get_row_value(self.structRow[3])
        dataDict: dict = {}
        self.structColLen = len(row5)
        for i in range(0, len(row5)):
            struct = DataStruct()
            struct._name = row5[i]
            struct._type = row6[i]
            struct._cs = row7[i]
            struct._def = row4[i]
            dataDict[i] = struct
        return dataDict

    def deal_single_sheet(self) -> None:
        """ 处理单张sheet """
        if not self.sheetStruct:
            return
        # print(sheet.max_row, sheet.max_column)
        # print('start to deal sheet: ', self.sheet.title)

        if self.sheetStruct.spcialType:
            self.deal_special_reach_row_data()
        else:
            self.deal_each_row_data()
        sheetindex.deal_sheet_index_file(self.sheetStruct.xlsxTitle, self.sheetStruct.sheetTitle,
                                         self.sheetStruct.clientName)

    def deal_special_reach_row_data(self) -> None:
        """ 处理特殊的导出格式，竖状 """
        dataStruct = self.get_data_struct()
        if not dataStruct:
            return

        totalJson = {}
        for row in range(self.startRow, self.sheet.max_row + 1):
            rowData: DataStruct = self.get_row_value(row)
            if len(rowData) == 0 or rowData[0] is None or 'C' not in rowData[2]:
                break

            totalJson[rowData[0]] = eachRowJson = {}
            col0: DataStruct = dataStruct[0]  # key
            eachRowJson[col0._name] = rowData[0]
            col3: DataStruct = dataStruct[3]  # value

            if rowData[1] == 'array':
                eachRowJson[col3._name] = str2list.str_to_list(rowData[3])
            elif rowData[1] == 'object':
                eachRowJson[col3._name] = json.loads(rowData[3])
            else:
                eachRowJson[col3._name] = rowData[3]

        self.deal_json_data(totalJson)

    def deal_each_row_data(self) -> None:
        """ 读取每行配置，处理导出数据 """
        dataStruct = self.get_data_struct()
        if not dataStruct:
            return

        # 判断是否有客户端或服务端字段
        haveClient = False
        haveServer = False
        for idx in range(0, len(dataStruct)):
            if haveClient is True and haveServer is True:
                break
            if 'C' in dataStruct[idx]._cs and haveClient is False:
                haveClient = True
            if 'S' in dataStruct[idx]._cs and haveServer is False:
                haveServer = True

        totalKey = self.sheetStruct.keyCount  # 表中配置的key数量
        maxRow = self.sheet.max_row

        totalJson = {}
        luaJson = {}  # 先生成临时json，再转化为lua
        for row in range(self.startRow, maxRow + 1):
            rowData = self.get_row_value(row)
            if len(rowData) == 0 or rowData[0] is None:
                break

            eachRowJson = totalJson
            eachLuaJson = luaJson
            for key in range(0, totalKey):
                if not eachRowJson.get(rowData[key]):
                    eachRowJson[rowData[key]] = {}
                    eachLuaJson[rowData[key]] = {}
                eachRowJson = eachRowJson.get(rowData[key])
                eachLuaJson = eachLuaJson.get(rowData[key])

            for col in range(0, self.structColLen):
                colStruct: DataStruct = dataStruct[col]
                if 'C' in colStruct._cs:
                    if colStruct._type == 'array':
                        eachRowJson[colStruct._name] = str2list.str_to_list(
                            rowData[col])
                    elif colStruct._type == 'object':
                        eachRowJson[colStruct._name] = json.loads(rowData[col])
                    else:
                        eachRowJson[colStruct._name] = rowData[col]
                if 'S' in colStruct._cs:
                    if colStruct._type == 'array':
                        eachLuaJson[colStruct._name] = str2list.str_to_list(
                            rowData[col])
                    elif colStruct._type == 'object':
                        eachLuaJson[colStruct._name] = json.loads(rowData[col])
                    else:
                        eachLuaJson[colStruct._name] = rowData[col]

        if haveClient:
            self.deal_json_data(totalJson)
        # else:
        #     print('\t\t【{0}】不需要导出json'.format(self.sheet.title))

        if haveServer:
            self.deal_lua_data(luaJson)
        # else:
        #     print("\t\t【{0}】不需要导出lua".format(self.sheet.title))

    def deal_json_data(self, obj: dict) -> None:
        """ 导出json数据 """
        struct = self.sheetStruct
        if not struct.clientName:
            # print(self.xlslUrl + ' --- 客户端配置名为空 -- 不导出json')
            return

        with open(self.outputRoot + "/" + struct.clientName, "w", encoding='utf-8') as outfile:
            json.dump(obj, outfile, indent=2, ensure_ascii=False)
            # outfile.write(json.dumps(obj, indent=4, ensure_ascii=True))

        self.deal_cfglist_json(struct.clientName)
        self.deal_config_ts(struct.clientName)

    def deal_lua_data(self, obj: dict) -> None:
        """ 导出lua数据 """
        struct = self.sheetStruct
        if not struct.serverName:
            # print(self.xlslUrl + ' --- 服务端配置名为空 -- 不导出lua')
            return

            # lua说明
        urlList = self.xlslUrl.split(os.sep)  # 获取路径中的xlsx文件名
        luaStr = "-- {0}\n-- {1}\n".format(urlList[-1], struct.serverName)
        luaStr += "-- %s\n\n" % time.strftime(
            "%Y-%m-%d %H:%M:%S", time.localtime())
        luaStr += "return"
        # json转化为lua
        wf = open(self.outputRoot + "/" + struct.serverName, 'w', 1, 'utf-8')
        wf.write(luaStr + json2lua.dic_to_lua_str(obj))
        wf.close()

    def deal_cfglist_json(self, client_name: str) -> None:
        """ 处理 cfglist.json 文件 """
        # print('start to export cfglist.json file')
        cfglistjson.deal_cfglist_json(client_name, self.outputRoot)
        # print('write cfglist.json successful!!!')

    def deal_config_ts(self, client_name: str) -> None:
        """ 导出config.d.ts文件 待处理 """
        # print('start to export config.ts file')
        dataDict: dict = self.get_data_struct()  # 传入结构体
        struct = configts.ConfigInterfaceStruct()
        struct.clientName = client_name
        struct.clientNameDef = self.sheet.title
        struct.dataObj = dataDict
        struct.outputRoot = self.outputRoot
        configts.deal_config_ts(struct)
