"""
把 excel 导出 json | lua
"""

import sys
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
import json
import str2list
import time
import json2lua
import os

class SheetStruct:
    """ 表的导出信息 """

    # 服务端导出文件名称
    def serverName(self):
        pass

    # 客户端导出文件名称
    def clientName(self):
        pass

    # key的数量
    def keyCount(self):
        pass

    # 特殊的格式
    def spcialType(self):
        pass


class DataStruct:
    """ 导出数据的结构体信息 """

    # 字段名
    def _name(self):
        pass

    # 类型 (number, string, array, object)
    def _type(self):
        pass

    # 导出字段（S服务端C客户端）
    def _cs(self):
        pass


class Excel2Json:

    xlslUrl = ''
    # 配表信息定义的行数
    structRow = [4, 5, 6, 7]
    # 配表真实数据要导出的列数，每行的后面可能有些说明，但是是不用导出，所以需要记录要导出的真实列数，也就是structRow行的真实列数
    structColLen = 0
    # 配表真实数据开始的行数
    startRow = 8
    sheet: worksheet.Worksheet = None
    # 表头结构体
    sheetStruct: SheetStruct = None
	# 所有json文件列表
    cfgListJson = "cfglist.json"

    def __init__(self, xlsxUrl: str) -> None:
        self.xlslUrl = xlsxUrl

    def readFile(self) -> None:
        wb = load_workbook(filename=self.xlslUrl)
        # print(wb.sheetnames)
        # print(wb.worksheets)
        for sheet in wb.worksheets:
            # title中以#开头的表示不导出
            if sheet.title[0] == '#':
                continue
            self.sheet = sheet
            if (not self.getSheetStruct()):
                continue
            self.dealSingleSheet()
        self.sheet = None
        self.xlslUrl = ''
        self.sheetStruct = None

    def getSheetStruct(self) -> SheetStruct:
        """ 获取当个sheet导出文件的名称以及key数量 """
        serverName = self.sheet.cell(row=1, column=2).value
        clientName = self.sheet.cell(row=1, column=5).value
        keyCount = self.sheet.cell(row=2, column=2).value
        specialType = self.sheet.cell(row=2, column=5).value
        if (not (serverName or clientName)):
            return None
        struct = SheetStruct()
        struct.serverName = serverName
        struct.clientName = clientName
        struct.keyCount = keyCount
        struct.spcialType = specialType == 1
        self.sheetStruct = struct
        return struct

    def getRowValue(self, row: int) -> list:
        """ 获取某行的数据 """
        columns = self.sheet.max_column
        rowData = []
        for i in range(1, columns+1):
            cellValue = self.sheet.cell(row=row, column=i).value
            # 单元格填0是需要导出的
            if cellValue == None:
                break
            rowData.append(cellValue)
        return rowData

    def getDataStruct(self) -> dict:
        """ 导出数据的结构体信息 """
        row4 = self.getRowValue(self.structRow[0])
        row5 = self.getRowValue(self.structRow[1])
        row6 = self.getRowValue(self.structRow[2])
        row7 = self.getRowValue(self.structRow[3])
        dataStruct = {}
        self.structColLen = len(row5)
        for i in range(0, len(row5)):
            struct = DataStruct()
            struct._name = row5[i]
            struct._type = row6[i]
            struct._cs = row7[i]
            dataStruct[i] = struct
        return dataStruct

    def dealSingleSheet(self) -> None:
        """ 处理单张sheet """
        if (not self.sheetStruct):
            return
        # print(sheet.max_row, sheet.max_column)
        print('开始处理sheet: ', self.sheet.title)

        if self.sheetStruct.spcialType:
            self.dealSpecailReachRowData()
        else:
            self.dealEachRowData()

    def dealSpecailReachRowData(self) -> None:
        """ 处理特殊的导出格式，竖状 """
        dataStruct = self.getDataStruct()
        if not dataStruct:
            return

        totalJson = {}
        for row in range(self.startRow, self.sheet.max_row + 1):
            rowData: DataStruct = self.getRowValue(row)
            if len(rowData) == 0 or rowData[0] == None or 'C' not in rowData[2]:
                break

            totalJson[rowData[0]] = eachRowJson = {}
            col0: DataStruct = dataStruct[0]  # key
            eachRowJson[col0._name] = rowData[0]
            col3: DataStruct = dataStruct[3]  # value

            if rowData[1] == 'array':
                eachRowJson[col3._name] = str2list.strToList(rowData[3])
            elif rowData[1] == 'object':
                eachRowJson[col3._name] = json.loads(rowData[3])
            else:
                eachRowJson[col3._name] = rowData[3]

        self.dealJsonData(totalJson)

    def dealEachRowData(self) -> None:
        """ 读取每行配置，处理导出数据 """
        dataStruct = self.getDataStruct()
        if not dataStruct:
            return

        # 判断是否有客户端或服务端字段
        haveClient = False
        haveServer = False
        for idx in range(0, len(dataStruct)):
            if haveClient == True and haveServer == True:
                break
            if 'C' in dataStruct[idx]._cs and haveClient == False:
                haveClient = True
            if 'S' in dataStruct[idx]._cs and haveServer == False:
                haveServer = True

        totalKey = self.sheetStruct.keyCount  # 表中配置的key数量
        maxRow = self.sheet.max_row

        totalJson = {}
        luaJson = {} # 先生成临时json，再转化为lua
        for row in range(self.startRow, maxRow+1):
            rowData = self.getRowValue(row)
            if len(rowData) == 0 or rowData[0] == None:
                break

            eachRowJson = totalJson
            eachLuaJson = luaJson
            for key in range(0, totalKey):
                if not eachRowJson.get(rowData[key]):
                    eachRowJson[rowData[key]] = {}
                    eachLuaJson[rowData[key]] = {}
                eachRowJson = eachRowJson.get(rowData[key])
                eachLuaJson = eachLuaJson.get(rowData[key])

            for col in range(0, self.structColLen):
                colStruct: DataStruct = dataStruct[col]
                if 'C' in colStruct._cs:
                    if colStruct._type == 'array':
                        eachRowJson[colStruct._name] = str2list.strToList(rowData[col])
                    elif colStruct._type == 'object':
                        eachRowJson[colStruct._name] = json.loads(rowData[col])
                    else:
                        eachRowJson[colStruct._name] = rowData[col]
                if 'S' in colStruct._cs:
                    if colStruct._type == 'array':
                        eachLuaJson[colStruct._name] = str2list.strToList(rowData[col])
                    elif colStruct._type == 'object':
                        eachLuaJson[colStruct._name] = json.loads(rowData[col])
                    else:
                        eachLuaJson[colStruct._name] = rowData[col]
        
        if haveClient:
            self.dealJsonData(totalJson)
        else:
            print('\t\t【{0}】不需要导出json'.format(self.sheet.title))

        if haveServer:
            self.dealLuaData(luaJson)
        else:
            print("\t\t【{0}】不需要导出lua".format(self.sheet.title))

    def dealJsonData(self, obj: dict) -> None:
        """ 导出json数据 """
        nameList = self.sheetStruct
        if not nameList.clientName:
            print(self.xlslUrl + ' --- 客户端配置名为空 -- 不导出json')
            return

        with open(outputRoot + "/" + nameList.clientName, "w", encoding='utf-8') as outfile:
            json.dump(obj, outfile, indent=2, ensure_ascii=False)
            # outfile.write(json.dumps(obj, indent=4, ensure_ascii=True))		
        
        self.dealCfglistJson(nameList.clientName)

    def dealLuaData(self, obj: dict) -> None:
        """ 导出lua数据 """
        nameList = self.sheetStruct
        if not nameList.serverName:
            print(self.xlslUrl + ' --- 服务端配置名为空 -- 不导出lua')
            return

		# lua说明
        urlList = self.xlslUrl.split(os.sep) # 获取路径中的xlsx文件名
        luaStr = "-- {0}\n-- {1}\n".format(urlList[-1], nameList.serverName)
        luaStr += "-- %s\n\n" % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        luaStr += "return"
		# json转化为lua
        wf = open(outputRoot + "/" + nameList.serverName, 'w', 1, 'utf-8')
        wf.write(luaStr + json2lua.dic_to_lua_str(obj))
        wf.close()

    def dealCfglistJson(self, clientName: str) -> None:
        """ 处理 cfglist.json 文件 """
        # print(clientName)
        cfglistjsonDir = os.path.join(outputRoot + "/" + self.cfgListJson)
        if not os.path.exists(cfglistjsonDir) or os.path.getsize(cfglistjsonDir) == 0:
            # 第一次写入cfglist.json文件，或者文件为空的。w写入方式
            with open(cfglistjsonDir, 'w', encoding='utf-8') as outfile:
                ary=[clientName] # json文件名数组
                json.dump(ary, outfile, indent=2, ensure_ascii=False)
        else:
            # cfglist.json文件存在
            newJsonAry = []
            isEmpty = os.path.getsize(cfglistjsonDir) == 0 # 判断文件大小，为0表示为空
            if isEmpty == True:
                print('cfglist.json文件是空的，有问题。请删除cfglist.json文件，重新导出')
                return
            with open(cfglistjsonDir, "r") as cfglistjson:
                # firstChar = cfglistjson.read(1) # 读取第一位字符判断，不存在就表示空文件，读取后记得要处理seek(0)
                # if not firstChar:
                #     print('cfglist.json文件是空的，有问题。请删除cfglist.json文件，重新导出')
                #     return
                # cfglistjson.seek(0) # 重置到文件头

                jsonAry: list = json.load(cfglistjson) # 这里若是文件空的，读入就会有问题 待处理
                if clientName not in jsonAry:
                    jsonAry.append(clientName) # 导出的json未存在，则加入
                    jsonAry.sort() # 排序
                    newJsonAry = jsonAry
            # 重新写入 （和上面一步整合）
            if newJsonAry and len(newJsonAry)> 0:
                with open(cfglistjsonDir, 'w') as cfglistjson:
                    json.dump(newJsonAry, cfglistjson, indent=2, ensure_ascii=False) 

    def dealConfigTs(self) -> None:
        """ 导出config.d.ts文件 待处理 """
        print('处理ts接口文件')

if __name__ == '__main__':
    # print(sys.argv)

    if sys.argv and len(sys.argv) > 2:
        # 倒入单独excel表导出
        if sys.argv and len(sys.argv) > 1 and sys.argv[1]:
            xlsxUrl = os.path.join(sys.argv[1])
        if sys.argv and len(sys.argv) > 2 and sys.argv[2]:
            outputRoot = os.path.join(sys.argv[2])
    else:
        # 运行所有的excel todo 测试
        currentpath = os.path.abspath(__file__)
        dirname = os.path.dirname(currentpath)
        xlsxUrl = os.path.normpath(os.path.join(dirname, "../test.xlsx"))
        outputRoot = os.path.normpath(os.path.join(dirname, "../output"))

    # print(xlsxUrl, outputRoot)

    # 若导出路径不存在，创建
    if not os.path.exists(outputRoot):
        os.makedirs(outputRoot)

    # 开始处理excel表
    excel2Json = Excel2Json(xlsxUrl)
    excel2Json.readFile()
